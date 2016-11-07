from openpyxl import load_workbook
wb = load_workbook(filename = 'xlsheet.xlsx')
sheet_ranges = wb['Sheet1']
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText

#dict = {'AN': 0, 'AP': 0, 'AR': 0, 'AS': 0, 'BR': 0, 'CG': 0, 'CH': 0, 'DD': 0, 'DL': 0, 'DN': 0, 'GA': 0, 'GJ': 0, 'HR': 0, 'HP': 0, 'JH': 0, 'JK': 0, 'KA': 0, 'KL': 0, 'LD': 0, 'MH': 0, 'ML': 0, 'MN': 0, 'MP': 0, 'MZ': 0, 'NL': 0, 'OD': 0, 'PB': 0, 'PY': 0, 'RJ': 0, 'SK': 0, 'TN': 0, 'TR': 0, 'TS': 0, 'UK': 0, 'UP': 0, 'WB': 0}; 
 
fromaddr = USERNAMEHERE

for x in xrange(1,4):
	a = str(x)
	#print(sheet_ranges[a].value)
	toaddr = sheet_ranges['A'+a].value
	msg = MIMEMultipart()
	msg['From'] = fromaddr
	msg['To'] = toaddr
	msg['Subject'] = "Spic-macay test mail"
	body = "Dear "+sheet_ranges['D'+a].value+" your registeration ID is "+sheet_ranges['C'+a].value
	msg.attach(MIMEText(body, 'plain'))
	try:
		server = smtplib.SMTP('smtp.gmail.com', 587)
		server.starttls()
		server.login(fromaddr, PASSWORDHERE)
		text = msg.as_string()
		server.sendmail(fromaddr, toaddr, text)
		server.quit()
		sheet_ranges['E'+a] = "SENT"
		ws = wb.active
		ws['E'+a] = "SENT"
		
		print("\n"+a+" sent")
	except:
		sheet_ranges['E'+a] = "FAILED"
		print("\n"+a+" not sent")
wb.save('xlsheet.xlsx')
